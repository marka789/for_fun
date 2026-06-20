#!/usr/bin/env python3
"""Generate hotel procurement Excel workbook with INDEX/MATCH, SUMIFS, no array formulas."""

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

LAST_ROW = 10001  # row 1 = header, rows 2-10001 = 10,000 data rows
DATA_START = 2
DATA_END = LAST_ROW


def set_headers(ws, headers):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def fill_formula_column(ws, col, formula_builder, start_row=DATA_START, end_row=DATA_END):
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=col, value=formula_builder(row))


def add_list_validation(ws, cell_range, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    dv.error = "Please select a value from the list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_backend(wb):
    ws = wb.create_sheet("Backend", 0)
    set_headers(
        ws,
        [
            "Item ID",
            "Brand",
            "Chinese Desc",
            "English Desc",
            "Red Threshold",
            "Lookup Key",
            "Item ID List",
            "Brand List",
            "Building List",
        ],
    )

    catalog = [
        ("APPLES", "FUJI", "富士苹果", "Fuji Apples", 20),
        ("APPLES", "GRANNY", "青苹果", "Granny Smith Apples", 20),
        ("DETERGENT", "TIDE", "汰渍洗衣液", "Tide Laundry Detergent", 10),
        ("DETERGENT", "LOCAL", "本地洗洁精", "Local Dish Soap", 15),
        ("TOWELS", "HOTEL", "酒店毛巾", "Hotel Bath Towels", 50),
        ("TOWELS", "HAND", "擦手巾", "Hand Towels", 30),
    ]
    for row_idx, row_data in enumerate(catalog, start=DATA_START):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.cell(row=row_idx, column=6, value=f'=A{row_idx}&"|"&B{row_idx}')

    fill_formula_column(ws, 6, lambda row: f'=A{row}&"|"&B{row}', start_row=DATA_START + len(catalog))

    item_ids = ["APPLES", "DETERGENT", "TOWELS"]
    brands = ["FUJI", "GRANNY", "TIDE", "LOCAL", "HOTEL", "HAND"]
    buildings = ["Building A", "Building B", "Tower 1", "Tower 2"]

    for idx, item_id in enumerate(item_ids, start=DATA_START):
        ws.cell(row=idx, column=7, value=item_id)
    for idx, brand in enumerate(brands, start=DATA_START):
        ws.cell(row=idx, column=8, value=brand)
    for idx, building in enumerate(buildings, start=DATA_START):
        ws.cell(row=idx, column=9, value=building)

    widths = {"A": 14, "B": 14, "C": 22, "D": 28, "E": 14, "F": 18, "G": 14, "H": 14, "I": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    return ws


def build_in_out_sheet(wb, title):
    ws = wb.create_sheet(title)
    set_headers(
        ws,
        [
            "Receipt Base",
            "Line",
            "Receipt No",
            "Item ID",
            "Brand",
            "Chinese Desc",
            "English Desc",
            "Date",
            "Qty",
            "Building",
            "Pic",
        ],
    )

    fill_formula_column(ws, 3, lambda row: f'=IF(A{row}="","",A{row}&"-"&TEXT(B{row},"00"))')
    fill_formula_column(
        ws,
        6,
        lambda row: (
            f'=IFERROR(INDEX(Backend!$C${DATA_START}:$C${DATA_END},'
            f'MATCH(D{row}&"|"&E{row},Backend!$F${DATA_START}:$F${DATA_END},0)),"")'
        ),
    )
    fill_formula_column(
        ws,
        7,
        lambda row: (
            f'=IFERROR(INDEX(Backend!$D${DATA_START}:$D${DATA_END},'
            f'MATCH(D{row}&"|"&E{row},Backend!$F${DATA_START}:$F${DATA_END},0)),"")'
        ),
    )

    add_list_validation(ws, f"D{DATA_START}:D{DATA_END}", "ItemID_List")
    add_list_validation(ws, f"E{DATA_START}:E{DATA_END}", "Brand_List")
    add_list_validation(ws, f"J{DATA_START}:J{DATA_END}", "Building_List")

    for col, width in zip("ABCDEFGHIJK", [16, 8, 18, 14, 14, 22, 28, 12, 8, 16, 20]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    return ws


def build_master(wb):
    ws = wb.create_sheet("Master")
    set_headers(
        ws,
        [
            "Receipt No",
            "Type",
            "Item ID",
            "Brand",
            "Chinese Desc",
            "English Desc",
            "Date",
            "Qty",
            "Building",
            "Balance (Item ID)",
            "Balance (Item+Brand)",
        ],
    )

    in_count = f"COUNTA(PurchaseIn!$A${DATA_START}:$A${DATA_END})"
    out_count = f"COUNTA(UseOut!$A${DATA_START}:$A${DATA_END})"

    def master_field(in_col, out_col, row):
        return (
            f'=IF(ROW()-1<={in_count},INDEX(PurchaseIn!${in_col}${DATA_START}:${in_col}${DATA_END},ROW()-1),'
            f'IF(ROW()-1-{in_count}<={out_count},INDEX(UseOut!${out_col}${DATA_START}:${out_col}${DATA_END},ROW()-1-{in_count}),""))'
        )

    fill_formula_column(ws, 1, lambda row: master_field("C", "C", row))
    fill_formula_column(ws, 2, lambda row: f'=IF(A{row}="","",IF(LEFT(A{row},2)="IN","IN","OUT"))')
    fill_formula_column(ws, 3, lambda row: master_field("D", "D", row))
    fill_formula_column(ws, 4, lambda row: master_field("E", "E", row))
    fill_formula_column(ws, 5, lambda row: master_field("F", "F", row))
    fill_formula_column(ws, 6, lambda row: master_field("G", "G", row))
    fill_formula_column(ws, 7, lambda row: master_field("H", "H", row))
    fill_formula_column(ws, 8, lambda row: master_field("I", "I", row))
    fill_formula_column(ws, 9, lambda row: master_field("J", "J", row))
    fill_formula_column(
        ws,
        10,
        lambda row: (
            f'=IF(C{row}="","",SUMIFS(PurchaseIn!$I${DATA_START}:$I${DATA_END},PurchaseIn!$D${DATA_START}:$D${DATA_END},C{row})-'
            f'SUMIFS(UseOut!$I${DATA_START}:$I${DATA_END},UseOut!$D${DATA_START}:$D${DATA_END},C{row}))'
        ),
    )
    fill_formula_column(
        ws,
        11,
        lambda row: (
            f'=IF(C{row}="","",SUMIFS(PurchaseIn!$I${DATA_START}:$I${DATA_END},PurchaseIn!$D${DATA_START}:$D${DATA_END},C{row},'
            f'PurchaseIn!$E${DATA_START}:$E${DATA_END},D{row})-'
            f'SUMIFS(UseOut!$I${DATA_START}:$I${DATA_END},UseOut!$D${DATA_START}:$D${DATA_END},C{row},'
            f'UseOut!$E${DATA_START}:$E${DATA_END},D{row}))'
        ),
    )

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        f"J{DATA_START}:J{DATA_END}",
        FormulaRule(
            formula=[
                f'AND(J{DATA_START}<>"",J{DATA_START}<IFERROR(INDEX(Backend!$E${DATA_START}:$E${DATA_END},MATCH(C{DATA_START},Backend!$G${DATA_START}:$G${DATA_END},0)),999999))'
            ],
            fill=red_fill,
        ),
    )
    ws.conditional_formatting.add(
        f"K{DATA_START}:K{DATA_END}",
        FormulaRule(
            formula=[
                f'AND(K{DATA_START}<>"",K{DATA_START}<IFERROR(INDEX(Backend!$E${DATA_START}:$E${DATA_END},MATCH(C{DATA_START}&"|"&D{DATA_START},Backend!$F${DATA_START}:$F${DATA_END},0)),999999))'
            ],
            fill=red_fill,
        ),
    )

    for col, width in zip("ABCDEFGHIJK", [18, 8, 14, 14, 22, 28, 12, 8, 16, 18, 20]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    return ws


def receipt_match_formula(result_col, row, search_cell="$C$1"):
    master_range = f"Master!${result_col}${DATA_START}:${result_col}${DATA_END}"
    receipt_range = f"Master!$A${DATA_START}:$A${DATA_END}"
    row_helper = f"$A{row}"
    return (
        f'=IFERROR(INDEX({master_range},AGGREGATE(15,6,ROW({receipt_range})/'
        f"(({receipt_range}={search_cell})+(LEFT({receipt_range},LEN({search_cell}))={search_cell})*({search_cell}<>\"\")),"
        f"{row_helper})),\"\")"
    )


def date_filter_formula(result_col, row):
    master_a = f"Master!$A${DATA_START}:$A${DATA_END}"
    master_c = f"Master!$C${DATA_START}:$C${DATA_END}"
    master_d = f"Master!$D${DATA_START}:$D${DATA_END}"
    master_g = f"Master!$G${DATA_START}:$G${DATA_END}"
    result_range = f"Master!${result_col}${DATA_START}:${result_col}${DATA_END}"
    row_helper = f"$A{row}"
    criteria = (
        f"({master_g}>=$C$28)*({master_g}<=$C$29)*($C$30=\"\")+"
        f"({master_c}=$C$30)*($C$30<>\"\")*({master_g}>=$C$28)*({master_g}<=$C$29)*($C$31=\"\")+"
        f"({master_c}=$C$30)*({master_d}=$C$31)*($C$30<>\"\")*($C$31<>\"\")*({master_g}>=$C$28)*({master_g}<=$C$29)"
    )
    return f'=IFERROR(INDEX({result_range},AGGREGATE(15,6,ROW({master_a})/({criteria}),{row_helper})),"")'


def build_search(wb):
    ws = wb.create_sheet("Search")
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

    ws["A1"] = "Receipt Search"
    ws["A1"].font = title_font
    ws["B1"] = "Search Receipt No"
    ws["B1"].font = Font(bold=True)

    headers_a = ["#", "Receipt No", "Type", "Item ID", "Brand", "Chinese", "English", "Date", "Qty"]
    for col, title in enumerate(headers_a, start=1):
        cell = ws.cell(row=3, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill

    for row in range(5, 25):
        ws.cell(row=row, column=1, value="=ROW()-4")

    result_cols = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for offset, col_letter in enumerate(result_cols[1:], start=0):
        for row in range(5, 25):
            ws.cell(row=row, column=2 + offset, value=receipt_match_formula(col_letter, row, search_cell="$C$1"))

    ws["A28"] = "Date / Item / Brand Search"
    ws["A28"].font = title_font
    for cell_ref, label in [
        ("B28", "Start Date"),
        ("B29", "End Date"),
        ("B30", "Item ID (optional)"),
        ("B31", "Brand (optional)"),
    ]:
        ws[cell_ref] = label
        ws[cell_ref].font = Font(bold=True)

    add_list_validation(ws, "C30", "ItemID_List")
    add_list_validation(ws, "C31", "Brand_List")

    for col, title in enumerate(headers_a, start=1):
        cell = ws.cell(row=33, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill

    for row in range(35, 55):
        ws.cell(row=row, column=1, value="=ROW()-34")

    for offset, col_letter in enumerate(result_cols[1:], start=0):
        for row in range(35, 55):
            ws.cell(row=row, column=2 + offset, value=date_filter_formula(col_letter, row))

    ws["A57"] = "Period Summary"
    ws["A57"].font = title_font
    ws["A58"] = "Total IN (period)"
    ws["A59"] = "Total OUT (period)"
    ws["A60"] = "Net (IN - OUT)"
    for row in range(58, 61):
        ws.cell(row=row, column=1).font = Font(bold=True)

    ws["B58"] = (
        f'=SUMIFS(PurchaseIn!$I${DATA_START}:$I${DATA_END},PurchaseIn!$H${DATA_START}:$H${DATA_END},">="&$C$28,'
        f'PurchaseIn!$H${DATA_START}:$H${DATA_END},"<="&$C$29)'
    )
    ws["B59"] = (
        f'=SUMIFS(UseOut!$I${DATA_START}:$I${DATA_END},UseOut!$H${DATA_START}:$H${DATA_END},">="&$C$28,'
        f'UseOut!$H${DATA_START}:$H${DATA_END},"<="&$C$29)'
    )
    ws["B60"] = "=B58-B59"

    ws["M63"] = "Inventory Snapshot"
    ws["M63"].font = title_font
    snap_headers = ["Item ID", "Brand", "Current Balance", "Red Threshold"]
    for col_offset, title in enumerate(snap_headers):
        cell = ws.cell(row=64, column=13 + col_offset, value=title)
        cell.font = header_font
        cell.fill = header_fill

    for row in range(65, 85):
        backend_row = row - 63
        ws.cell(row=row, column=13, value=f"=Backend!A{backend_row}")
        ws.cell(row=row, column=14, value=f"=Backend!B{backend_row}")
        ws.cell(
            row=row,
            column=15,
            value=(
                f'=IF(M{row}="","",SUMIFS(PurchaseIn!$I${DATA_START}:$I${DATA_END},PurchaseIn!$D${DATA_START}:$D${DATA_END},M{row},'
                f'PurchaseIn!$E${DATA_START}:$E${DATA_END},N{row})-'
                f'SUMIFS(UseOut!$I${DATA_START}:$I${DATA_END},UseOut!$D${DATA_START}:$D${DATA_END},M{row},'
                f'UseOut!$E${DATA_START}:$E${DATA_END},N{row}))'
            ),
        )
        ws.cell(
            row=row,
            column=16,
            value=(
                f'=IF(M{row}="","",IFERROR(INDEX(Backend!$E${DATA_START}:$E${DATA_END},'
                f'MATCH(M{row}&"|"&N{row},Backend!$F${DATA_START}:$F${DATA_END},0)),""))'
            ),
        )

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "O65:O84",
        FormulaRule(formula=['AND(O65<>"",P65<>"",O65<P65)'], fill=red_fill),
    )

    for col, width in zip("ABCDEFGHI", [6, 18, 12, 14, 14, 22, 28, 12, 8]):
        ws.column_dimensions[col].width = width
    for col, width in zip("MNOP", [14, 14, 16, 14]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"
    return ws


def add_named_ranges(wb):
    ranges = {
        "ItemID_List": f"Backend!$G${DATA_START}:$G${DATA_END}",
        "Brand_List": f"Backend!$H${DATA_START}:$H${DATA_END}",
        "Building_List": f"Backend!$I${DATA_START}:$I${DATA_END}",
    }
    for name, ref in ranges.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_backend(wb)
    build_in_out_sheet(wb, "PurchaseIn")
    build_in_out_sheet(wb, "UseOut")
    build_master(wb)
    build_search(wb)
    add_named_ranges(wb)

    output_path = "/workspace/procurement_workbook.xlsx"
    wb.save(output_path)
    print(f"Created {output_path}")


if __name__ == "__main__":
    main()
